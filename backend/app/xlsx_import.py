"""Leitor da planilha `ERP - Ecommer.xlsx`.

Constrói um Snapshot em memória a partir das abas de dados (Produtos, Entradas,
Kits, Composicao Kits, Vendas, Ads, Configuracoes). Usado por:
- o importador de seed (grava numa org demo)
- os testes golden (valida o engine contra os valores da planilha)

IDs sintéticos são atribuídos na ordem das linhas (a ordem das vendas = ordem das
linhas = cronológica, o que preserva a semântica FIFO da planilha).
"""
from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import date, datetime

import openpyxl

from app.services.snapshot import (
    SAd,
    SKit,
    SKitComponent,
    SLot,
    SProduct,
    SSale,
    Snapshot,
)

DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "data", "ERP-Ecommer.xlsx")


@dataclass
class ParsedSettings:
    taxa_shopee_pct: float = 0.20
    taxa_fixa: float = 4.0
    taxa_afiliado_pct: float = 0.0


@dataclass
class Parsed:
    snapshot: Snapshot
    settings: ParsedSettings
    org_name: str = "Minha Loja"


def _as_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _num(v, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def parse_workbook(path: str = DEFAULT_XLSX) -> Parsed:
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Configuracoes ---
    cfg = wb["Configuracoes"]
    settings = ParsedSettings(
        taxa_shopee_pct=_num(cfg["B4"].value, 0.20),
        taxa_fixa=_num(cfg["B5"].value, 4.0),
        taxa_afiliado_pct=_num(cfg["B6"].value, 0.0),
    )

    # --- Produtos --- (dados a partir da linha 4)
    products: list[SProduct] = []
    sku_to_pid: dict[str, int] = {}
    ws = wb["Produtos"]
    for r in range(4, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        if not sku:
            continue
        sku = str(sku).strip()
        pid = len(products) + 1
        sku_to_pid[sku] = pid
        ativo = str(ws.cell(r, 5).value or "Sim").strip().lower() in ("sim", "true", "1", "yes")
        products.append(
            SProduct(
                id=pid,
                sku=sku,
                nome=str(ws.cell(r, 2).value or sku).strip(),
                variacao=(str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else None),
                dropdown_name=str(ws.cell(r, 4).value or sku).strip(),
                ativo=ativo,
            )
        )

    # --- Entradas (lotes FIFO) ---
    lots: list[SLot] = []
    ws = wb["Entradas"]
    for r in range(4, ws.max_row + 1):
        sku = ws.cell(r, 3).value
        d = _as_date(ws.cell(r, 2).value)
        if not sku or d is None:
            continue
        sku = str(sku).strip()
        pid = sku_to_pid.get(sku)
        if pid is None:
            continue
        lots.append(
            SLot(
                id=len(lots) + 1,
                product_id=pid,
                data_entrada=d,
                qty_in=int(_num(ws.cell(r, 5).value, 0)),
                unit_cost=_num(ws.cell(r, 6).value, 0.0),
                lote_code=(str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else None),
            )
        )

    # --- Kits ---
    kits: list[SKit] = []
    sku_to_kid: dict[str, int] = {}
    ws = wb["Kits"]
    for r in range(4, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        if not sku:
            continue
        sku = str(sku).strip()
        kid = len(kits) + 1
        sku_to_kid[sku] = kid
        ativo = str(ws.cell(r, 3).value or "Sim").strip().lower() in ("sim", "true", "1", "yes")
        kits.append(
            SKit(
                id=kid,
                sku=sku,
                nome=str(ws.cell(r, 2).value or sku).strip(),
                components=[],
                ativo=ativo,
                preco_referencia=(_num(ws.cell(r, 7).value) if ws.cell(r, 7).value else None),
            )
        )
    kit_by_id = {k.id: k for k in kits}

    # --- Composicao Kits (BOM) ---
    ws = wb["Composicao Kits"]
    for r in range(4, ws.max_row + 1):
        kit_sku = ws.cell(r, 1).value
        prod_sku = ws.cell(r, 3).value
        if not kit_sku or not prod_sku:
            continue
        kid = sku_to_kid.get(str(kit_sku).strip())
        pid = sku_to_pid.get(str(prod_sku).strip())
        if kid is None or pid is None:
            continue
        kit_by_id[kid].components.append(
            SKitComponent(product_id=pid, qty=int(_num(ws.cell(r, 4).value, 1)))
        )

    # --- Vendas ---
    sales: list[SSale] = []
    ws = wb["Vendas"]
    for r in range(4, ws.max_row + 1):
        d = _as_date(ws.cell(r, 1).value)
        sku = ws.cell(r, 4).value  # coluna D = SKU resolvido (valor em cache)
        if d is None or not sku:
            continue
        sku = str(sku).strip()
        pid = sku_to_pid.get(sku)
        kid = sku_to_kid.get(sku)
        if pid is None and kid is None:
            continue
        item_type = "product" if pid is not None else "kit"
        qty = int(_num(ws.cell(r, 5).value, 0))
        preco = _num(ws.cell(r, 6).value, 0.0)
        shopee_pct = _num(ws.cell(r, 8).value, settings.taxa_shopee_pct)
        taxa_fixa = _num(ws.cell(r, 10).value, settings.taxa_fixa)
        afiliado_pct = _num(ws.cell(r, 11).value, 0.0)

        # "Outras Taxas": normalmente a coluna M. Mas a planilha permite sobrescrever a
        # Receita Líquida (col N) manualmente — quando isso acontece, a diferença é uma
        # dedução extra real do pedido (cupom/taxa avulsa). Recuperamos esse valor a
        # partir do líquido gravado para reproduzir os números exatamente.
        receita = qty * preco
        n_val = ws.cell(r, 14).value
        if isinstance(n_val, (int, float)):
            outras = receita - receita * shopee_pct - taxa_fixa - receita * afiliado_pct - n_val
            if abs(outras) < 1e-9:
                outras = 0.0
        else:
            outras = _num(ws.cell(r, 13).value, 0.0)

        sales.append(
            SSale(
                id=len(sales) + 1,
                data_venda=d,
                item_type=item_type,
                qty=qty,
                preco_unitario=preco,
                taxa_shopee_pct=shopee_pct,
                taxa_fixa=taxa_fixa,
                taxa_afiliado_pct=afiliado_pct,
                outras_taxas=outras,
                product_id=pid,
                kit_id=kid,
                pedido=(str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else None),
            )
        )

    # --- Ads ---
    ads: list[SAd] = []
    ws = wb["Ads"]
    for r in range(4, ws.max_row + 1):
        d = _as_date(ws.cell(r, 1).value)
        valor = ws.cell(r, 3).value
        if d is None or valor is None or valor == "":
            continue
        ads.append(
            SAd(
                id=len(ads) + 1,
                data=d,
                valor=_num(valor, 0.0),
                canal=(str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else None),
            )
        )

    snap = Snapshot(products=products, lots=lots, kits=kits, sales=sales, ads=ads)
    return Parsed(snapshot=snap, settings=settings)
